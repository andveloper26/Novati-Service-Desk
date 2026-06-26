from openpyxl import load_workbook
from datetime import datetime

def obtener_siguiente_orden():

    libro = load_workbook("datos/ordenes.xlsx")

    hoja = libro["Ordenes"]

    ultima_fila = hoja.max_row

    if ultima_fila == 1:
        return 53037

    ultimo_numero = hoja.cell(row=ultima_fila, column=1).value

    return int(ultimo_numero) + 1

def guardar_orden(
    numero,
    fecha,
    hora,
    cliente,
    celular,
    equipo,
    marca,
    modelo,
    falla,
    estado,
    observacion
):
    libro = load_workbook("datos/ordenes.xlsx")

    hoja = libro["Ordenes"]

    hoja.append([
        numero,
        fecha,
        hora,
        cliente,
        celular,
        equipo,
        marca,
        modelo,
        falla,
        estado,
        observacion
    ])

    libro.save("datos/ordenes.xlsx")

def obtener_ordenes():

    libro = load_workbook("datos/ordenes.xlsx")

    hoja = libro["Ordenes"]

    ordenes = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        ordenes.append(fila)

    return ordenes

def buscar_orden(numero):

    ordenes = obtener_ordenes()

    print("Buscando:", numero)

    for orden in ordenes:

        print("Orden encontrada:", orden[0], type(orden[0]))

        if str(orden[0]).strip() == str(numero).strip():

            return orden

    return None

def actualizar_estado(numero_orden, nuevo_estado):

    libro = load_workbook("datos/ordenes.xlsx")

    hoja = libro["Ordenes"]

    for fila in range(2, hoja.max_row + 1):

        if hoja.cell(row=fila, column=1).value == numero_orden:

            hoja.cell(row=fila, column=10).value = nuevo_estado

            break

    libro.save("datos/ordenes.xlsx")


def guardar_historial(numero_orden, estado, observacion, tecnico):

    libro = load_workbook("datos/historial.xlsx")

    hoja = libro["Historial_Ordenes"]

    ultimo_id = hoja.max_row

    ahora = datetime.now()

    hoja.append([
        ultimo_id,
        numero_orden,
        ahora.strftime("%d/%m/%Y"),
        ahora.strftime("%H:%M"),
        estado,
        observacion,
        tecnico
    ])

    libro.save("datos/historial.xlsx")

def obtener_turnos():

    libro = load_workbook("datos/turnos.xlsx")

    hoja = libro["Turnos"]

    turnos = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        turnos.append(fila)

    return turnos



def convertir_turno_a_orden(turno):

    numero = obtener_siguiente_orden()

    ahora = datetime.now()

    fecha = ahora.strftime("%d/%m/%Y")

    hora = ahora.strftime("%H:%M")

    guardar_orden(
        numero,
        fecha,
        hora,
        turno[3],      # Cliente
        turno[4],      # Celular
        turno[5],      # Tipo de Equipo
        turno[6],      # Marca
        turno[7],      # Modelo
        turno[8],      # Falla
        "Ingresado",
        ""
    )

    eliminar_turno(turno[0])

def eliminar_turno(id_turno):

    libro = load_workbook("datos/turnos.xlsx")

    hoja = libro["Turnos"]

    for fila in range(2, hoja.max_row + 1):

        if hoja.cell(row=fila, column=1).value == id_turno:

            hoja.delete_rows(fila)

            break

    libro.save("datos/turnos.xlsx")

def obtener_estadisticas():

    ordenes = obtener_ordenes()
    turnos = obtener_turnos()

    total_ordenes = len(ordenes)
    total_turnos = len(turnos)

    listos = 0
    conflictos = 0

    for orden in ordenes:

        if orden[9] == "Listo para entregar":
            listos += 1

        if orden[9] == "En conflicto":
            conflictos += 1

    return total_turnos, total_ordenes, listos, conflictos

def guardar_turno(
    cliente,
    celular,
    equipo,
    marca,
    modelo,
    falla
):

    libro = load_workbook("datos/turnos.xlsx")

    hoja = libro["Turnos"]

    nuevo_id = hoja.max_row

    ahora = datetime.now()

    fecha = ahora.strftime("%d/%m/%Y")

    hora = ahora.strftime("%H:%M")

    hoja.append([
        nuevo_id,
        fecha,
        hora,
        cliente,
        celular,
        equipo,
        marca,
        modelo,
        falla,
        "Pendiente"
    ])

    libro.save("datos/turnos.xlsx")
